import openpyxl
import os
import logging
import traceback
from PIL import Image, ImageDraw, ImageFont

# Paths
angel_file_path = "angel.xlsx"
photos_folder = "photos"  # All files (design images) are in this single folder

def read_excel(file_path):
    """
    Reads the angel.xlsx file and returns a list of SKU numbers and their quantities.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        log_error(f"Error: Excel file '{file_path}' not found.")
        return [], []

    sheet = wb.active
    skus = []
    qtys = []

    # Read the SKU numbers and quantities from the Excel file
    for sku_cell, qty_cell in sheet.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        if sku_cell and qty_cell:
            skus.append(str(sku_cell))  # Use SKU exactly as it appears, ensure it's a string
            qtys.append(int(qty_cell))  # Add corresponding quantity to the list
    
    return skus, qtys

def filter_skus_by_type(skus, sku_type):
    """
    Filters the SKUs based on the selected type (TS or HD).
    """
    return [sku for sku in skus if sku.split('-')[1] == sku_type]

def get_user_selection():
    """
    Asks the user to select TS or HD, then color and size.
    """
    sku_type = input("Do you want to display 'TS' or 'HD' SKUs? (Enter TS or HD): ").strip().upper()

    # Ensure the user chooses either TS or HD
    while sku_type not in ['TS', 'HD']:
        print("Invalid input! Please enter 'TS' or 'HD'.")
        sku_type = input("Do you want to display 'TS' or 'HD' SKUs? (Enter TS or HD): ").strip().upper()

    # Ask for color and size choices
    color = input("Choose a color (BLK, GRN, WHT): ").strip().upper()
    while color not in ['BLK', 'GRN', 'WHT']:
        print("Invalid input! Please choose from 'BLK', 'GRN', or 'WHT'.")
        color = input("Choose a color (BLK, GRN, WHT): ").strip().upper()

    size = input("Choose a size (S, M, L, XL): ").strip().upper()
    while size not in ['S', 'M', 'L', 'XL']:
        print("Invalid input! Please choose from 'S', 'M', 'L', or 'XL'.")
        size = input("Choose a size (S, M, L, XL): ").strip().upper()

    return sku_type, color, size

def check_images(skus, qtys, sku_type, color, size):
    """
    Checks if the image files corresponding to the selected SKUs exist and displays them with quantity overlay.
    Marks the corresponding quantity cell in yellow in the Excel file.
    """
    for sku, qty in zip(skus, qtys):
        sku_parts = sku.split('-')
        
        # Extract the SKU number (the part between TS/HD and NAME)
        if sku_parts[1] != sku_type:  # Ensure we are filtering the correct type (TS or HD)
            continue
        
        sku_number = sku_parts[2]  # This is the SKU number between TS/HD and NAME (e.g., "1", "2", "3", etc.)

        # Check if the color and size match
        if sku_parts[4] == color and sku_parts[5] == size:
            image_path = os.path.join(photos_folder, f"{sku_number}.png")  # e.g., "1.png"

            if not os.path.exists(image_path):
                log_error(f"FILE: {sku_number}.png could not be found.")
                continue

            print(f"Found image for SKU {sku} with quantity {qty}")

            # Create image with text overlay
            create_image_with_overlay(image_path, qty)

            # Mark the corresponding quantity cell in yellow after the image is processed
            mark_qty_cell_yellow(sku, qty)

def create_image_with_overlay(image_path, qty):
    """
    Creates a new image with the quantity overlayed as text with a larger font size.
    """
    try:
        # Open the image
        image = Image.open(image_path)
        print(f"Opened image: {image_path}")

        # Create a drawing context
        draw = ImageDraw.Draw(image)

        # Define the text to be overlayed
        overlay_text = f"Qty: {qty}"

        # Load a custom font and increase its size (1000% bigger)
        font_path = "arial.ttf"  # Path to your font file (you can use another font)
        base_font_size = 40  # Base font size, you can tweak this based on your image size

        # Scale the font size by 1000%
        scaled_font_size = base_font_size * 10  # 1000% = 10 times larger

        try:
            font = ImageFont.truetype(font_path, scaled_font_size)  # Load the font with the new size
        except IOError:
            # Fallback to the default font if the custom font is not found
            print(f"Font {font_path} not found. Using default font.")
            font = ImageFont.load_default()

        # Calculate the text width and height using textbbox (Bounding box of the text)
        text_bbox = draw.textbbox((0, 0), overlay_text, font=font)
        text_width = text_bbox[2] - text_bbox[0]  # Width of the text (x2 - x1)
        text_height = text_bbox[3] - text_bbox[1]  # Height of the text (y2 - y1)

        # Get the dimensions of the image
        width, height = image.size

        # Calculate the position to center the text
        text_position = ((width - text_width) // 2, height - text_height - 10)

        # Add the text to the image
        draw.text(text_position, overlay_text, font=font, fill="white")

        # Save the new image as a temporary file
        temp_image_path = os.path.join(photos_folder, f"temp_{os.path.basename(image_path)}")
        image.save(temp_image_path)

        print(f"Created image with overlay: {temp_image_path}")
        
        # Optionally, show the image (requires an image viewer, or use image.show())
        image.show()

    except Exception as e:
        log_error(f"Error creating image with overlay for {image_path}: {e}")

def mark_qty_cell_yellow(sku, qty):
    """
    Marks the corresponding cell in the Excel sheet yellow for the found SKU.
    """
    try:
        wb = openpyxl.load_workbook(angel_file_path)
        sheet = wb.active

        # Find the row of the SKU
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=2):
            if str(row[0].value) == sku:
                # Mark the corresponding quantity cell yellow
                row[1].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                print(f"Marked quantity for SKU {sku} in yellow.")
                break
        
        wb.save(angel_file_path)
    except Exception as e:
        log_error(f"Error marking yellow cell for SKU {sku}: {e}")

def log_error(message):
    """
    Logs error messages to a file.
    """
    logging.basicConfig(filename="error_log.txt", level=logging.ERROR)
    logging.error(f"{message}\n{traceback.format_exc()}")

def main():
    """
    Main function that processes the Excel file and handles the user's SKU type, color, and size selection.
    """
    try:
        # Read SKUs and quantities from Excel
        skus, qtys = read_excel(angel_file_path)
        if not skus:
            print("No data found in the selected Excel file.")
            return
        
        # Ask for user input to filter by SKU type (TS/HD), color, and size
        sku_type, color, size = get_user_selection()

        # Filter the SKUs by the selected type (TS or HD)
        filtered_skus = filter_skus_by_type(skus, sku_type)
        filtered_qtys = [qtys[skus.index(sku)] for sku in filtered_skus]

        # Check and display images for the selected color and size
        check_images(filtered_skus, filtered_qtys, sku_type, color, size)
        
        print("Processing complete!")

    except Exception as e:
        log_error(f"Unexpected error: {e}")
        print("An unexpected error occurred. Check the log file for details.")

if __name__ == "__main__":
    main()
